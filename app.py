# give openpyxl an alias to make it shorter
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# define a function with a file
def process_workbook(filename):
# use the load_workbook function to upload the excel file
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # figure out how many rows your sheet has using sheet.max_row
    # iterate over all of the rows

    for row in range(2, sheet.max_row + 1): # the + 1 makes sure that it includes all of the rows
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # Select the cells in row 2 to 4
    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    # add values in column 4 and create a chart with these values
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    # save the workbook in a new file so that you don't overwrite the original file in case your code has a bug
    wb.save(filename)
